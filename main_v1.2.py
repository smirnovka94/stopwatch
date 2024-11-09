import tkinter as tk
import time
import json

import openpyxl
from datetime import datetime
import psutil


class StopwatchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Секундомер")
        self.root.wm_attributes("-topmost", 1)
        self.is_running = False
        self.start_time = 0
        self.elapsed_time = 0
        self.index = 0
        self.time_without_work = {}  # Список для хранения времени без работы
        self.comments = []

        self.comment_var = tk.StringVar()

        self.create_widgets()

    def create_widgets(self):
        self.time_display = tk.Label(self.root, font=("Helvetica", 48), text="00:00:00")
        self.time_display.pack()

        self.comment_entry = tk.Entry(self.root, textvariable=self.comment_var, width=50)
        self.comment_entry.pack()

        self.start_button = tk.Button(self.root, text="Старт", command=self.start)
        self.start_button.pack(side='left')

        self.pause_button = tk.Button(self.root, text="Пауза", command=self.pause, state='disabled')
        self.pause_button.pack(side='left')

        self.stop_button = tk.Button(self.root, text="Сохранить", command=self.stop)
        self.stop_button.pack(side='left')

        self.update_display()

    def update_display(self):
        if self.is_running:
            self.elapsed_time = time.time() - self.start_time
            self.time_display.config(text=self.format_time(self.elapsed_time))
        self.root.after(1000, self.update_display)  # Обновляем время каждую секунду

    def format_time(self, seconds):
        hours, remainder = divmod(int(seconds), 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours:02}:{minutes:02}:{seconds:02}"

    def start(self):
        if self.comment_var.get() and not self.is_running:
            self.start_time = time.time() - self.elapsed_time
            self.time_without_work[self.index] = [self.comment_var.get(), datetime.now(), None, None, None]

            self.record_event('Старт')
            self.is_running = True
            self.start_button.config(state='disabled')
            self.pause_button.config(state='normal')


    def pause(self):
        if self.is_running:
            self.record_event('Пауза')
            time_stop = datetime.now()
            time_delta = time_stop - self.time_without_work[self.index][1]
            total_seconds = int(time_delta.total_seconds())
            formatted_time = f"{total_seconds // 3600}:{(total_seconds % 3600) // 60}:{total_seconds % 60}"

            self.time_without_work[self.index][1] = self.time_without_work[self.index][1].strftime("%H:%M:%S")
            self.time_without_work[self.index][2] = time_stop.strftime("%H:%M:%S")
            self.time_without_work[self.index][3] = formatted_time
            self.index += 1
            self.is_running = False
            self.start_button.config(state='normal')
            self.pause_button.config(state='disabled')
            self.comment_var.set("")  # Сброс комментария



    def stop(self):
        self.is_running = False
        self.record_event('Стоп')
        self.save_to_json()
        self.save_to_excel()



    @staticmethod
    def gather_system_info():
        # Получаем информацию о использовании ЦП
        cpu_usage = psutil.cpu_percent(interval=1)

        # Получаем информацию о использовании памяти
        memory_info = psutil.virtual_memory()
        memory_usage = memory_info.percent

        # Получаем список активных процессов
        active_processes = [proc.info for proc in psutil.process_iter(['pid', 'name', 'cpu_percent'])]

        # Формируем данные для записи в JSON
        system_info = {
            "cpu_usage_percent": cpu_usage,
            "memory_usage_percent": memory_usage,
            "active_processes": active_processes
        }

        return system_info

    def record_event(self, button_type):
        system_info = self.gather_system_info()
        timestamp = datetime.now().isoformat("#", "seconds"),
        time = datetime.now().strftime("%H:%M:%S")

        event = {
            'comment': self.comment_var.get() if button_type in ['Старт', 'Пауза'] else "",
            'timestamp': timestamp,
            'time': time,
            'cpu_usage_percent': system_info['cpu_usage_percent'],
            'memory_usage_percent': system_info['memory_usage_percent'],
            'active_processes': system_info['active_processes'],
            'type': button_type,
        }
        self.comments.append(event)

    def save_to_json(self):
        filename = f"stopwatch_data_{datetime.now().strftime('%Y.%m.%d_%H%M%S')}.json"
        with open(filename, 'w') as json_file:
            json.dump(self.comments, json_file, ensure_ascii=False, indent=4)

    def save_to_excel(self):
        filename = f"stopwatch_data_{datetime.now().strftime('%Y.%m.%d_%H%M%S')}.xlsx"

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        dict_data = self.time_without_work
        sheet.cell(row=1, column=1, value="Комментарий")
        sheet.cell(row=1, column=2, value="Время старта")
        sheet.cell(row=1, column=3, value="Время завершения")
        sheet.cell(row=1, column=4, value="Продолжение")

        for i, item in enumerate(dict_data.keys()):
            sheet.cell(row=i + 2, column=1, value=dict_data[item][0])
            sheet.cell(row=i + 2, column=2, value=dict_data[item][1])
            sheet.cell(row=i + 2, column=3, value=dict_data[item][2])
            sheet.cell(row=i + 2, column=4, value=dict_data[item][3])

        workbook.save(filename)


if __name__ == "__main__":
    root = tk.Tk()
    app = StopwatchApp(root)
    root.mainloop()
